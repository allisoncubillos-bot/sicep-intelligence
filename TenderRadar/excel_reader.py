"""Lectura del Excel de cantidades a texto, para alimentar a Claude.

El Excel ("Anexos 3 y 4. Cantidades requeridas y oferta") trae las cantidades
por producto, año y mes. En vez de parsear celdas específicas (la estructura
varía por agente), volcamos las hojas a texto tabular y dejamos que Claude lo
interprete.
"""
import logging

from openpyxl import load_workbook

log = logging.getLogger(__name__)

MAX_ROWS_PER_SHEET = 300
MAX_CHARS = 40_000


def excel_to_text(path: str | None) -> str:
    """Vuelca el contenido del Excel a texto tabular (hojas, filas no vacías)."""
    if not path:
        return ""
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:  # noqa: BLE001 - archivos .xls antiguos u otros
        log.warning("No se pudo leer el Excel '%s': %s", path, exc)
        return ""

    parts: list[str] = []
    for ws in wb.worksheets:
        parts.append(f"### Hoja: {ws.title}")
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i >= MAX_ROWS_PER_SHEET:
                parts.append("... (filas adicionales omitidas)")
                break
            cells = ["" if v is None else str(v) for v in row]
            if any(c.strip() for c in cells):
                parts.append("\t".join(cells).rstrip())
    wb.close()
    return "\n".join(parts)[:MAX_CHARS]
